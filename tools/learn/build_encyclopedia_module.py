"""Football Learning Engine -- the full Encyclopedia module (READS FULL
ENCYCLOPEDIA operation).

Extends the earlier Defensive Coverages module (build_coverage_module.py,
still the sole owner of the 32 COVER_*/coverage-family FB_CONCEPT nodes and
all learn_lessons/learn_exercises) into a complete football encyclopedia
spanning every domain the source workbook actually supports: positions,
personnel, formations, route tree, passing concepts, run game, blocking,
pass protection, quarterback play, defensive fronts/personnel/pressures,
special teams, situational football, play calling, scouting, history,
geometry, coaching, officiating, rules, offensive systems, plus
team-specific NFL/CFB scheme profiles and historical statistical leaders.

--- SOURCE ---
Reads directly from the user's workbook at SOURCE_XLSX (not committed to
this repo -- same convention as the other large source .xlsx files already
sitting untracked alongside it) via openpyxl, at import time. This is
DELIBERATE: copying ~2,500 rows of real, already-well-structured source text
into a Python literal would be a transcription exercise, not an import --
reading the live file means this script is genuinely re-runnable if the
workbook is revised, and every string in the database traces back to an
exact (sheet, row) the same way the coverage module traces to workbook rows.

--- SCHEMA (see module docstring further down at CONCEPT records for the
why) ---
Reuses `knowledge_nodes`/`knowledge_edges` (the same safe-to-extend,
zero-other-consumers tables the coverage module already established) with
THREE node_types:
  FB_CONCEPT              -- a teachable concept (this module's node_type,
                              same as every coverage concept -- one flat
                              universe, distinguished by a `domain` field in
                              payload_json, not by a proliferation of types).
  FB_TEAM_SCHEME_PROFILE   -- an NFL/FBS team's 2026 scheme profile. Kept as
                              its OWN node_type, deliberately never mixed
                              into FB_CONCEPT, because the source itself
                              labels every one of these "Staff verified.
                              Scheme components are lineage- and prior-
                              film-based PROJECTIONS; 2026 usage/rates
                              require a multi-game film sample" -- this is
                              team-specific, unconfirmed-by-film content,
                              structurally distinct from a general, durable
                              football concept (see NEEDS_VERIFICATION
                              handling below).
  FB_HISTORICAL_RECORD     -- a real, source-cited statistical-leader season
                              (Greatest Offenses/Defenses) -- a fact about a
                              specific team-season, not a teachable concept.

Every node's payload_json always carries: `domain` (the top-level Learn
category, see LEARN_DOMAINS below), `subcategory` (the sheet's own grouping
column, preserved verbatim), `fields` (a dict of the SHEET'S OWN real
columns, snake_cased -- deliberately NOT a fixed universal template; a route
has route-shaped fields, a position has position-shaped fields, a scheme
profile has 29 real columns -- see the FULL FOOTBALL LEARNING ENCYCLOPEDIA
mission's explicit "different concept types should support different
appropriate fields, do not fabricate fields just to fill a template"), and
`source_rows` (list of {sheet, row} -- full provenance to this exact
workbook, for every field, always).

--- MERGING (Section 2 of the mission: "no unexplained ignored content,
duplicates get MERGED not duplicated") ---
Several sheets describe the SAME real concept from different angles (e.g.
"Inside Zone" appears in both Scheme Concepts, with fan-facing "what it
is/why/how answered/film tells" framing, and Run Game & Blocking, with
coaching-facing "how it works/coaching point/defensive answer" framing).
Where a real, confirmed overlap exists (checked by hand against both
sheets' actual content, not assumed from name similarity alone -- see
MERGE_MAP below), this script produces ONE node with BOTH sheets' fields
merged under clearly separate keys and BOTH sheets' rows in source_rows --
never two duplicate nodes for the same concept, and never a silent
overwrite that drops one source's real content. Concepts that only
resemble each other by name but are NOT the same thing per the source
(e.g. Scheme Concepts treats "Flood" and "Sail" as two distinct pass
concepts; Passing Concepts' "Flood / Sail" row is kept as a single
supplementary annotation attached to BOTH, never forced into declaring
them identical) are kept distinct, honoring the source's own distinctions.

--- COVERAGE MODULE INTEGRATION ---
Scheme Concepts' 9 coverage rows (Cover 0 through Bracket/Cone) and Defense
Masterclass's ~13 coverage rows describe concepts that ALREADY EXIST as
FB_CONCEPT nodes from build_coverage_module.py (COVER_0, COVER_1_ROBBER,
COVER_2, TAMPA_2, COVER_3, COVER_4, COVER_6, PALMS_COVERAGE,
BRACKET_COVERAGE, ROBBER_COVERAGE, MATCH_COVERAGE). This script ENRICHES
those existing nodes additively (new `film_room_tells`/`offensive_counter`
keys folded into their existing payload_json, via the exact same
ON CONFLICT DO UPDATE upsert pattern, never replacing what
build_coverage_module.py already established) rather than creating
parallel duplicate coverage nodes. It never touches learn_lessons/
learn_exercises -- those remain entirely build_coverage_module.py's.

--- NEEDS_VERIFICATION HANDLING (Section 11) ---
NFL/FBS Scheme Profiles are real, richly detailed, and genuinely useful --
but every row's own text says its scheme components are lineage/prior-film
projections, not confirmed 2026 film evidence. `verification_status` for
every FB_TEAM_SCHEME_PROFILE node is set to `LINEAGE_PROJECTED_NEEDS_FILM_
VERIFICATION` (never SOURCE_BACKED, which this codebase's whole session has
used to mean "independently confirmed," not "coach's staff is confirmed but
scheme is a lineage guess"). The export/frontend must show this distinction
plainly, never silently presenting a projection as a confirmed fact.

--- WHAT THIS SCRIPT DOES NOT DO ---
Does not touch Grid, Creator/feasibility, Trivia, or any Engine refresh
script -- purely additive new content in knowledge_nodes/knowledge_edges,
read by a new export script and a new frontend encyclopedia UI only.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(REPO_ROOT))
from tools.quiz_export import engine  # noqa: E402

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required (pip install openpyxl)", file=sys.stderr)
    raise

SOURCE_XLSX = Path(
    "/Users/enterprise2/Desktop/2026 NFL Draft Guide/"
    "Reads_Football_Encyclopedia_Leak_Safe_700_Question_Master.xlsx"
)
SOURCE_ID = "FOOTBALL_ENCYCLOPEDIA_LEAK_SAFE_700_MASTER_WORKBOOK"
SOURCE_BACKED = "SOURCE_BACKED"
LINEAGE_PROJECTED = "LINEAGE_PROJECTED_NEEDS_FILM_VERIFICATION"

# The Learn encyclopedia's top-level taxonomy -- one entry per browsable
# category, matching the mission's requested structure. `order` sets the
# beginner->advanced browse order; `tier` is informational (used in the
# coverage report, not by the UI).
LEARN_DOMAINS = [
    {"id": "FOOTBALL_101", "label": "Football 101", "order": 1},
    {"id": "RULES", "label": "Rules & Officiating", "order": 2},
    {"id": "POSITIONS", "label": "Positions & Responsibilities", "order": 3},
    {"id": "PERSONNEL", "label": "Personnel Groupings", "order": 4},
    {"id": "FORMATIONS", "label": "Formations & Alignments", "order": 5},
    {"id": "ROUTE_TREE", "label": "Route Tree", "order": 6},
    {"id": "PASSING_CONCEPTS", "label": "Passing Concepts", "order": 7},
    {"id": "RUN_GAME", "label": "Run Game", "order": 8},
    {"id": "BLOCKING", "label": "Offensive Line & Blocking", "order": 9},
    {"id": "PASS_PROTECTION", "label": "Pass Protection", "order": 10},
    {"id": "QB_PLAY", "label": "Quarterback Play", "order": 11},
    {"id": "DEFENSIVE_FRONTS", "label": "Defensive Fronts", "order": 12},
    {"id": "DEFENSIVE_PERSONNEL", "label": "Defensive Personnel", "order": 13},
    {"id": "RUN_FITS", "label": "Run Fits", "order": 14},
    {"id": "COVERAGES", "label": "Coverages", "order": 15},
    {"id": "PRESSURES", "label": "Pressures & Blitzes", "order": 16},
    {"id": "DEFENSIVE_PHILOSOPHY", "label": "Defensive Structure & Philosophy", "order": 17},
    {"id": "OFFENSIVE_SYSTEMS", "label": "Offensive Systems", "order": 18},
    {"id": "SPECIAL_TEAMS", "label": "Special Teams", "order": 19},
    {"id": "SITUATIONAL", "label": "Situational Football", "order": 20},
    {"id": "PLAY_CALLING", "label": "Play Calling Language", "order": 21},
    {"id": "GEOMETRY", "label": "Football Geometry & Leverage", "order": 22},
    {"id": "COACHING", "label": "Coaching & Game Planning", "order": 23},
    {"id": "SCOUTING", "label": "Scouting & Analytics", "order": 24},
    {"id": "FILM_STUDY", "label": "Film Study", "order": 25},
    {"id": "HISTORY", "label": "Football History & Evolution", "order": 26},
    {"id": "NFL_SCHEMES", "label": "NFL Offensive & Defensive Schemes", "order": 27},
    {"id": "CFB_SCHEMES", "label": "College Offensive & Defensive Schemes", "order": 28},
    {"id": "GREAT_UNITS", "label": "Great Offenses & Great Defenses", "order": 29},
]


def _slug(text: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "_", str(text).upper()).strip("_")


def _snake(text: str) -> str:
    text = re.sub(r"[/\-]", " ", str(text))
    text = re.sub(r"[^a-zA-Z0-9 ]", "", text)
    return re.sub(r"\s+", "_", text.strip().lower())


def _node_id(canonical_id: str) -> str:
    return f"KN|FB_CONCEPT|{canonical_id}"


def _team_node_id(canonical_id: str) -> str:
    return f"KN|FB_TEAM_SCHEME_PROFILE|{canonical_id}"


def _hist_node_id(canonical_id: str) -> str:
    return f"KN|FB_HISTORICAL_RECORD|{canonical_id}"


def _load_wb():
    if not SOURCE_XLSX.exists():
        raise SystemExit(f"ABORT: source workbook not found at {SOURCE_XLSX}")
    return openpyxl.load_workbook(SOURCE_XLSX, data_only=True)


def _rows(ws, min_row=2):
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out = []
    for i, row in enumerate(ws.iter_rows(min_row=min_row, values_only=True), start=min_row):
        if all(v is None for v in row):
            continue
        out.append((i, dict(zip(header, row))))
    return out


def _ensure_schema(c) -> None:
    c.execute(
        "INSERT INTO sources(source_id, source_name, attribution_required, approved_for_import, notes) "
        "VALUES (?,?,0,1,?) ON CONFLICT(source_id) DO NOTHING",
        (SOURCE_ID, "Reads Football Encyclopedia (Leak-Safe 700 Question Master workbook)",
         "Full multi-sheet edition -- positions, personnel, formations, route tree, passing "
         "concepts, run game, blocking, protection, QB play, defensive fronts/personnel/"
         "pressures, special teams, situational football, play calling, scouting, history, "
         "geometry, coaching, officiating, rules, offensive systems, NFL/CFB team scheme "
         "profiles, and historical statistical leaders."),
    )


def _upsert_concept(c, canonical_id, label, domain, subcategory, fields, source_rows,
                     verification_status=SOURCE_BACKED, extra_payload=None):
    node_id = _node_id(canonical_id)
    payload = {"domain": domain, "subcategory": subcategory, "fields": fields,
               "source_rows": source_rows}
    if extra_payload:
        payload.update(extra_payload)
    c.execute(
        "INSERT INTO knowledge_nodes(node_id, node_type, canonical_id, label, competition_id, "
        "payload_json, verification_status) VALUES (?,?,?,?,?,?,?) "
        "ON CONFLICT(node_type, canonical_id) DO UPDATE SET "
        "label=excluded.label, payload_json=excluded.payload_json, verification_status=excluded.verification_status",
        (node_id, "FB_CONCEPT", canonical_id, label, None, json.dumps(payload), verification_status),
    )
    return node_id


def _enrich_existing_concept(c, canonical_id, new_fields, source_rows):
    """Additive-only enrichment of an EXISTING FB_CONCEPT node
    (build_coverage_module.py's coverage concepts) -- reads the current
    payload_json, adds new keys under `encyclopedia_fields` (never
    overwrites an existing key the coverage module already set), appends
    source_rows, writes back. No-op (returns False) if the node doesn't
    exist -- callers must have gotten the canonical_id right, this never
    creates a new node as a side effect of a typo."""
    row = c.execute(
        "SELECT payload_json FROM knowledge_nodes WHERE node_type='FB_CONCEPT' AND canonical_id=?",
        (canonical_id,),
    ).fetchone()
    if not row:
        return False
    payload = json.loads(row["payload_json"])
    enc = payload.setdefault("encyclopedia_fields", {})
    for k, v in new_fields.items():
        enc.setdefault(k, v)
    existing_rows = payload.setdefault("encyclopedia_source_rows", [])
    existing_rows.extend(r for r in source_rows if r not in existing_rows)
    c.execute(
        "UPDATE knowledge_nodes SET payload_json=? WHERE node_type='FB_CONCEPT' AND canonical_id=?",
        (json.dumps(payload), canonical_id),
    )
    return True


def _add_edge(c, source_canonical, predicate, target_canonical, node_type="FB_CONCEPT",
              confidence=1.0, verification_status=SOURCE_BACKED):
    prefix = "KN|" + node_type + "|"
    src = prefix + source_canonical
    tgt = prefix + target_canonical
    edge_id = f"KE|{src}|{predicate}|{tgt}"
    c.execute(
        "INSERT OR IGNORE INTO knowledge_edges(edge_id, source_node_id, predicate, target_node_id, "
        "season_start, season_end, source_id, verification_status, confidence, payload_json) "
        "VALUES (?,?,?,?,?,?,?,?,?,?)",
        (edge_id, src, predicate, tgt, None, None, SOURCE_ID, verification_status, confidence, "{}"),
    )


# ============================================================================
# Generic ingestion for "simple" sheets: one row = one concept, a name
# column, an optional subcategory column, and every other real column
# becomes a `fields` entry verbatim (snake_cased key, untouched value) --
# no template, no invented fields, exactly what the sheet itself provides.
# ============================================================================

# Rules & Game Mechanics' own "Category" values that are genuinely
# Football-101-level basics (how downs/clock/scoring/the field work) rather
# than rules/officiating detail -- routed to FOOTBALL_101 instead of RULES
# (see the mission's explicit Football 101 section: "downs and distance,
# scoring, possession, clock... field layout"). Everything else in that
# sheet (Scrimmage, Passing, Ball security, Kicking, Turnovers, Penalties,
# Administration, Strategy) stays in RULES.
FOOTBALL_101_CATEGORIES = {"Game structure", "Clock", "Scoring", "Field"}

SIMPLE_SHEETS = [
    # (sheet_name, name_col, subcat_col, domain, id_prefix, dedupe_within_sheet)
    ("Positions & Techniques", "Position / Role", "Unit", "POSITIONS", "POSITION"),
    ("Rules & Game Mechanics", "Rule / Concept", "Category", "RULES", "RULE"),
    ("Officiating & Penalties", "Rule Concept", "Phase", "RULES", "PENALTY"),
    ("Special Teams Masterclass", "Concept", "Phase", "SPECIAL_TEAMS", "SPECIALTEAMS"),
    ("Situational Football", "Situation", "Situation Group", "SITUATIONAL", "SITUATION"),
    ("Play Calling Language", "Term / System", "Layer", "PLAY_CALLING", "PLAYCALL"),
    ("Scouting Analytics Roster", "Concept / Metric", "Domain", "SCOUTING", "SCOUTMETRIC"),
    ("Football History & Evolution", "Innovation / Topic", "Era / Lens", "HISTORY", "HISTORY"),
    ("Quarterback Masterclass", "Skill / Read", "Phase", "QB_PLAY", "QBSKILL"),
    ("Football Geometry & Leverage", "Concept", "Domain", "GEOMETRY", "GEOMETRY"),
    ("Coaching & Game Planning", "Concept", "Stage", "COACHING", "COACHING"),
    ("How To Read Schemes", "Film-Study Action", "Step", "FILM_STUDY", "FILMSTUDY"),
    ("Film Charting Data Dictionary", "Field", "Data Type", "FILM_STUDY", "CHARTFIELD"),
    ("Pass Protection", "Protection / Technique", "Layer", "PASS_PROTECTION", "PROTECTION"),
]


def ingest_simple_sheets(c, wb, report: dict):
    for sheet_name, name_col, subcat_col, domain, prefix in SIMPLE_SHEETS:
        ws = wb[sheet_name]
        rows = _rows(ws)
        header = [c2.value for c2 in next(ws.iter_rows(min_row=1, max_row=1))]
        seen_ids: dict = {}
        imported = 0
        for row_num, rec in rows:
            name = rec.get(name_col)
            if not name:
                continue
            canonical_id = f"{prefix}_{_slug(name)}"
            # A handful of names repeat across Phase/Step groupings (e.g. the
            # same "Rule Concept" name never happens here, but guard anyway) --
            # disambiguate deterministically rather than silently overwrite.
            if canonical_id in seen_ids:
                seen_ids[canonical_id] += 1
                canonical_id = f"{canonical_id}_{seen_ids[canonical_id]}"
            else:
                seen_ids[canonical_id] = 1
            fields = {
                _snake(col): val for col, val in rec.items()
                if col not in (name_col,) and val is not None and str(val).strip()
            }
            subcategory = rec.get(subcat_col) if subcat_col else None
            row_domain = domain
            if sheet_name == "Rules & Game Mechanics" and subcategory in FOOTBALL_101_CATEGORIES:
                row_domain = "FOOTBALL_101"
            _upsert_concept(
                c, canonical_id, str(name), row_domain, subcategory, fields,
                [{"sheet": sheet_name, "row": row_num}],
            )
            imported += 1
        report[sheet_name] = {
            "data_rows": len(rows), "imported": imported, "domain": domain,
        }


# ============================================================================
# Route Tree Encyclopedia -- its own handler (skips the "Numbering warning"
# meta-row, keeps the tree-number field, preserves the source's own honesty
# disclaimers about historical player associations verbatim).
# ============================================================================

def ingest_route_tree(c, wb, report: dict):
    ws = wb["Route Tree Encyclopedia"]
    rows = _rows(ws)
    imported = 0
    for row_num, rec in rows:
        route = rec.get("Route")
        if not route or route == "Numbering warning":
            continue
        canonical_id = f"ROUTE_{_slug(route)}"
        fields = {
            _snake(col): val for col, val in rec.items()
            if col not in ("Route",) and val is not None and str(val).strip()
        }
        _upsert_concept(
            c, canonical_id, str(route), "ROUTE_TREE", rec.get("Family"), fields,
            [{"sheet": "Route Tree Encyclopedia", "row": row_num}],
        )
        imported += 1
    report["Route Tree Encyclopedia"] = {"data_rows": len(rows), "imported": imported, "domain": "ROUTE_TREE"}


# ============================================================================
# Scheme Concepts -- split by (Side, Family) into the right domain, since
# this one sheet spans run concepts, pass concepts, offensive systems,
# defensive fronts, defensive personnel, coverages (merged into the
# EXISTING coverage module, never duplicated), pressures, and defensive
# philosophy families. Every real MERGE decision below was made by reading
# both sheets' actual row content (see module docstring), not by name
# matching alone.
# ============================================================================

SCHEME_CONCEPT_FAMILY_DOMAIN = {
    "Run scheme": "RUN_GAME", "Gap run": "RUN_GAME", "Perimeter run": "RUN_GAME",
    "Zone complement": "RUN_GAME", "Play-action family": "RUN_GAME", "Run-pass option": "RUN_GAME",
    "Option run": "RUN_GAME", "Option family": "RUN_GAME",
    "Passing system": "OFFENSIVE_SYSTEMS", "Offensive family": "OFFENSIVE_SYSTEMS",
    "Pass concept": "PASSING_CONCEPTS",
    "Front": "DEFENSIVE_FRONTS", "Front family": "DEFENSIVE_FRONTS", "Front technique": "DEFENSIVE_FRONTS",
    "Personnel/front": "DEFENSIVE_PERSONNEL", "Personnel": "DEFENSIVE_PERSONNEL",
    "Coverage": "COVERAGES",
    "Pressure family": "PRESSURES", "Pressure presentation": "PRESSURES",
    "Defensive family": "DEFENSIVE_PHILOSOPHY",
}

# Existing build_coverage_module.py canonical_ids these Scheme Concepts rows
# enrich (never duplicate). Anything NOT in this map with Family=="Coverage"
# becomes a new FB_CONCEPT instead (there is one: none currently -- every
# Scheme Concepts coverage row maps to an existing node).
SCHEME_CONCEPTS_COVERAGE_ENRICH = {
    "Cover 0": "COVER_0", "Cover 1 Robber": "COVER_1_ROBBER", "Cover 2 Zone": "COVER_2",
    "Tampa 2": "TAMPA_2", "Cover 3 Match": "COVER_3", "Cover 4 Quarters": "COVER_4",
    "Cover 6": "COVER_6", "Palms / 2-Read": "PALMS_COVERAGE", "Bracket / Cone": "BRACKET_COVERAGE",
}

# Concept-name merge targets in Run Game & Blocking / Passing Concepts
# (case-insensitive exact match on the OTHER sheet's own concept name).
RUN_MERGE_TARGETS = {
    "wide zone": "wide zone", "inside zone": "inside zone", "duo": "duo", "power": "power",
    "counter gt": "counter gt", "pin-pull sweep": "pin-pull", "mid zone": "mid zone",
    "zone read": "zone read",
}
PASS_MERGE_TARGETS = {
    "mesh": "mesh", "four verticals": "four verticals", "y-cross": "y-cross",
    "drive": "drive", "dagger": "dagger", "smash": "smash", "spacing": "spacing", "stick": "stick",
}
# Flood/Sail: Passing Concepts' single "Flood / Sail" row applies to BOTH
# distinct Scheme Concepts entries (see module docstring -- the source
# itself distinguishes Flood from Sail; this is real shared enrichment, not
# a forced merge of the two into one).
FLOOD_SAIL_ROW_NAME = "flood / sail"


def _scheme_concept_canonical(family: str, name: str) -> str:
    fam_prefix = {
        "Run scheme": "RUN", "Gap run": "RUN", "Perimeter run": "RUN", "Zone complement": "RUN",
        "Play-action family": "RUN", "Run-pass option": "RUN", "Option run": "RUN", "Option family": "RUN",
        "Passing system": "SYSTEM", "Offensive family": "SYSTEM", "Pass concept": "PASSCONCEPT",
        "Front": "FRONT", "Front family": "FRONT", "Front technique": "FRONT",
        "Personnel/front": "DEFPERSONNEL", "Personnel": "DEFPERSONNEL",
        "Pressure family": "PRESSURE", "Pressure presentation": "PRESSURE",
        "Defensive family": "DEFPHILOSOPHY",
    }.get(family, "CONCEPT")
    return f"{fam_prefix}_{_slug(name)}"


def ingest_scheme_concepts(c, wb, report: dict):
    ws = wb["Scheme Concepts"]
    rows = _rows(ws)
    run_ws = wb["Run Game & Blocking"]
    run_by_name = {str(rec.get("Concept / Technique", "")).strip().lower(): (rn, rec)
                   for rn, rec in _rows(run_ws)}
    pass_ws = wb["Passing Concepts"]
    pass_by_name = {str(rec.get("Concept", "")).strip().lower(): (rn, rec)
                     for rn, rec in _rows(pass_ws)}
    flood_sail_row = pass_by_name.get(FLOOD_SAIL_ROW_NAME)

    imported = 0
    enriched_coverage = 0
    for row_num, rec in rows:
        name = rec.get("Concept")
        family = rec.get("Family")
        side = rec.get("Side")
        if not name or not family:
            continue

        if family == "Coverage":
            target = SCHEME_CONCEPTS_COVERAGE_ENRICH.get(name)
            if target:
                ok = _enrich_existing_concept(
                    c, target,
                    {
                        "why_coaches_use_it": rec.get("Why coaches use it"),
                        "how_opponents_answer_it": rec.get("How opponents answer it"),
                        "film_room_tells": rec.get("Film-room tells"),
                    },
                    [{"sheet": "Scheme Concepts", "row": row_num}],
                )
                if ok:
                    enriched_coverage += 1
                    continue
            # Fall through to a normal new node if no mapping exists (none
            # currently, kept as a safety net so a future workbook edit
            # never silently drops a coverage row).

        domain = SCHEME_CONCEPT_FAMILY_DOMAIN.get(family, "OFFENSIVE_SYSTEMS" if side == "Offense" else "DEFENSIVE_PHILOSOPHY")
        canonical_id = _scheme_concept_canonical(family, name)
        fields = {
            "what_it_is": rec.get("What it is"),
            "why_coaches_use_it": rec.get("Why coaches use it"),
            "how_opponents_answer_it": rec.get("How opponents answer it"),
            "film_room_tells": rec.get("Film-room tells"),
        }
        source_rows = [{"sheet": "Scheme Concepts", "row": row_num}]
        name_key = name.strip().lower()

        merge_target = None
        if domain == "RUN_GAME" and name_key in RUN_MERGE_TARGETS:
            merge_target = run_by_name.get(RUN_MERGE_TARGETS[name_key])
            if merge_target:
                mrow, mrec = merge_target
                fields["how_it_works"] = mrec.get("How It Works")
                fields["offensive_coaching_point"] = mrec.get("Offensive Coaching Point")
                fields["defensive_answer"] = mrec.get("Defensive Answer")
                source_rows.append({"sheet": "Run Game & Blocking", "row": mrow})
        elif domain == "PASSING_CONCEPTS" and name_key in PASS_MERGE_TARGETS:
            merge_target = pass_by_name.get(PASS_MERGE_TARGETS[name_key])
            if merge_target:
                mrow, mrec = merge_target
                fields["route_distribution"] = mrec.get("Route Distribution / Idea")
                fields["quarterback_read"] = mrec.get("Quarterback Read")
                fields["defensive_answer"] = mrec.get("Defensive Answer")
                source_rows.append({"sheet": "Passing Concepts", "row": mrow})
        elif domain == "PASSING_CONCEPTS" and name_key in ("flood", "sail") and flood_sail_row:
            mrow, mrec = flood_sail_row
            fields["route_distribution"] = mrec.get("Route Distribution / Idea")
            fields["quarterback_read"] = mrec.get("Quarterback Read")
            fields["defensive_answer"] = mrec.get("Defensive Answer")
            source_rows.append({"sheet": "Passing Concepts", "row": mrow})

        fields = {k: v for k, v in fields.items() if v is not None and str(v).strip()}
        _upsert_concept(c, canonical_id, name, domain, family, fields, source_rows)
        imported += 1

    report["Scheme Concepts"] = {
        "data_rows": len(rows), "imported": imported, "enriched_existing_coverage": enriched_coverage,
        "domain": "multiple (run/pass/systems/fronts/personnel/pressure/philosophy + coverage enrichment)",
    }


# ============================================================================
# Run Game & Blocking -- only the rows NOT already merged into Scheme
# Concepts above (new run concepts, blocking techniques, run fits).
# ============================================================================

RUN_GAME_FAMILY_DOMAIN = {
    "Run families": "RUN_GAME", "Option": "RUN_GAME",
    "Blocking rules": "BLOCKING", "Back technique": "RUN_GAME", "Run defense": "RUN_FITS",
}
RUN_GAME_PREFIX = {
    "Run families": "RUN", "Option": "RUN", "Blocking rules": "BLOCKTECH",
    "Back technique": "RUNTECH", "Run defense": "RUNFIT",
}
_ALREADY_MERGED_RUN_NAMES = set(RUN_MERGE_TARGETS.values())


def ingest_run_game_remainder(c, wb, report: dict):
    ws = wb["Run Game & Blocking"]
    rows = _rows(ws)
    imported = 0
    for row_num, rec in rows:
        name = rec.get("Concept / Technique")
        family = rec.get("Family")
        if not name or not family:
            continue
        if family == "Run families" and name.strip().lower() in _ALREADY_MERGED_RUN_NAMES:
            continue  # already merged into the matching Scheme Concepts node
        domain = RUN_GAME_FAMILY_DOMAIN.get(family, "RUN_GAME")
        prefix = RUN_GAME_PREFIX.get(family, "RUN")
        canonical_id = f"{prefix}_{_slug(name)}"
        fields = {
            "how_it_works": rec.get("How It Works"),
            "offensive_coaching_point": rec.get("Offensive Coaching Point"),
            "defensive_answer": rec.get("Defensive Answer"),
        }
        fields = {k: v for k, v in fields.items() if v is not None and str(v).strip()}
        _upsert_concept(c, canonical_id, name, domain, family, fields,
                         [{"sheet": "Run Game & Blocking", "row": row_num}])
        imported += 1
    report["Run Game & Blocking"] = {
        "data_rows": len(rows), "imported": imported,
        "merged_into_scheme_concepts": len(_ALREADY_MERGED_RUN_NAMES),
        "domain": "RUN_GAME / BLOCKING / RUN_FITS",
    }


# ============================================================================
# Passing Concepts -- only the rows NOT already merged into Scheme Concepts.
# ============================================================================

PASSING_CONCEPTS_DOMAIN = {
    "Quick game": "PASSING_CONCEPTS", "Man beaters": "PASSING_CONCEPTS", "Three-level": "PASSING_CONCEPTS",
    "Vertical stretch": "PASSING_CONCEPTS", "Crossing": "PASSING_CONCEPTS",
    "Two-high beaters": "PASSING_CONCEPTS", "Coverage manipulation": "PASSING_CONCEPTS",
    "Screens": "PASSING_CONCEPTS", "Movement": "PASSING_CONCEPTS", "RPO": "RUN_GAME",
    "Progression": "QB_PLAY",
}
_ALREADY_MERGED_PASS_NAMES = set(PASS_MERGE_TARGETS.values()) | {FLOOD_SAIL_ROW_NAME}


def ingest_passing_concepts_remainder(c, wb, report: dict):
    ws = wb["Passing Concepts"]
    rows = _rows(ws)
    imported = 0
    for row_num, rec in rows:
        name = rec.get("Concept")
        family = rec.get("Family")
        if not name or not family:
            continue
        if name.strip().lower() in _ALREADY_MERGED_PASS_NAMES:
            continue
        domain = PASSING_CONCEPTS_DOMAIN.get(family, "PASSING_CONCEPTS")
        prefix = {"RPO": "RPO", "Progression": "QBREAD"}.get(family, "PASSCONCEPT")
        canonical_id = f"{prefix}_{_slug(name)}"
        fields = {
            "route_distribution": rec.get("Route Distribution / Idea"),
            "quarterback_read": rec.get("Quarterback Read"),
            "defensive_answer": rec.get("Defensive Answer"),
        }
        fields = {k: v for k, v in fields.items() if v is not None and str(v).strip()}
        _upsert_concept(c, canonical_id, name, domain, family, fields,
                         [{"sheet": "Passing Concepts", "row": row_num}])
        imported += 1
    report["Passing Concepts"] = {
        "data_rows": len(rows), "imported": imported,
        "merged_into_scheme_concepts": len(_ALREADY_MERGED_PASS_NAMES),
        "domain": "PASSING_CONCEPTS / RUN_GAME(RPO) / QB_PLAY(progression)",
    }


# ============================================================================
# Trench Play Masterclass -- per-Unit domain routing (this sheet mixes
# offensive-line technique, defensive-line technique, front mechanics and
# grading philosophy under one column set).
# ============================================================================

TRENCH_UNIT_DOMAIN = {
    "Offensive line": "BLOCKING", "Defensive line": "DEFENSIVE_FRONTS",
    "Front mechanics": "DEFENSIVE_FRONTS", "Evaluation": "FILM_STUDY",
}


def ingest_trench_play(c, wb, report: dict):
    ws = wb["Trench Play Masterclass"]
    rows = _rows(ws)
    imported = 0
    for row_num, rec in rows:
        name = rec.get("Technique")
        unit = rec.get("Unit")
        if not name or not unit:
            continue
        domain = TRENCH_UNIT_DOMAIN.get(unit, "BLOCKING")
        canonical_id = f"TECH_{_slug(name)}"
        fields = {
            "definition": rec.get("Definition"), "film_evidence": rec.get("Film Evidence"),
            "football_value": rec.get("Football Value"), "app_tags": rec.get("App Tags"),
        }
        fields = {k: v for k, v in fields.items() if v is not None and str(v).strip()}
        _upsert_concept(c, canonical_id, name, domain, unit, fields,
                         [{"sheet": "Trench Play Masterclass", "row": row_num}])
        imported += 1
    report["Trench Play Masterclass"] = {
        "data_rows": len(rows), "imported": imported, "domain": "BLOCKING / DEFENSIVE_FRONTS / FILM_STUDY",
    }


# ============================================================================
# Formation Reading Lab -- three real sections: (1) an 8-step pre-snap read
# ORDER (-> FILM_STUDY), (2) offensive formations (-> FORMATIONS), (3) a
# defensive read-order + front-family section that OVERLAPS Scheme Concepts'
# front rows (merged, not duplicated) and defensive personnel packages
# (merged with Scheme Concepts' personnel rows).
# ============================================================================

FORMATION_LAB_FRONT_MERGE = {
    # Formation Reading Lab name (lowercase) -> Scheme Concepts canonical_id
    # already created by ingest_scheme_concepts() above. Scheme Concepts has
    # no standalone "Over"/"Under"/"Bear" rows -- only "4-3 Over"/"4-3
    # Under"/"Bear Front" -- so these merge into THOSE canonical_ids.
    "over": "FRONT_4_3_OVER", "under": "FRONT_4_3_UNDER", "bear": "FRONT_BEAR_FRONT",
}
FORMATION_LAB_PERSONNEL_MERGE = {
    "nickel / 4-2-5": "DEFPERSONNEL_NICKEL_4_2_5", "3-3-5 stack": "DEFPERSONNEL_3_3_5_STACK",
    "dime / dollar": "DEFPERSONNEL_DIME",
}


def ingest_formation_reading_lab(c, wb, report: dict):
    ws = wb["Formation Reading Lab"]
    rows = _rows(ws)
    imported = 0
    enriched = 0
    for row_num, rec in rows:
        phase = rec.get("Phase")
        name = rec.get("Formation / Diagnostic")
        if not name or not phase:
            continue
        name_key = name.strip().lower()
        fields_common = {
            "what_to_identify": rec.get("What To Identify"),
            "what_it_usually_tells_you": rec.get("What It Usually Tells You"),
            "confirmation_cue": rec.get("Confirmation Cue"),
            "terminology_variation_warning": rec.get("Terminology / Variation Warning"),
            "common_misread": rec.get("Common Misread"),
        }
        fields_common = {k: v for k, v in fields_common.items() if v and str(v).strip()}
        source_rows = [{"sheet": "Formation Reading Lab", "row": row_num}]

        if phase == "Read order":
            canonical_id = f"FILMSTUDY_{_slug(name)}"
            _upsert_concept(c, canonical_id, name, "FILM_STUDY", "Pre-snap read order", fields_common, source_rows)
            imported += 1
        elif phase in ("Core offense", "Special offense"):
            canonical_id = f"FORMATION_{_slug(name)}"
            _upsert_concept(c, canonical_id, name, "FORMATIONS", phase, fields_common, source_rows)
            imported += 1
        elif phase == "Defense read order":
            canonical_id = f"FILMSTUDY_DEF_{_slug(name)}"
            _upsert_concept(c, canonical_id, name, "FILM_STUDY", "Defensive pre-snap read order", fields_common, source_rows)
            imported += 1
        elif phase == "Base defense":
            canonical_id = f"FRONT_BASE_{_slug(name)}"
            _upsert_concept(c, canonical_id, name, "DEFENSIVE_FRONTS", "Base personnel front", fields_common, source_rows)
            imported += 1
        elif phase == "Sub defense":
            target = FORMATION_LAB_PERSONNEL_MERGE.get(name_key)
            if target and _enrich_existing_concept(c, target, fields_common, source_rows):
                enriched += 1
            else:
                canonical_id = f"DEFPERSONNEL_{_slug(name)}"
                _upsert_concept(c, canonical_id, name, "DEFENSIVE_PERSONNEL", "Sub-package personnel", fields_common, source_rows)
                imported += 1
        elif phase == "Front family":
            target = FORMATION_LAB_FRONT_MERGE.get(name_key)
            if target and _enrich_existing_concept(c, target, fields_common, source_rows):
                enriched += 1
            elif name_key == "tite / mint":
                # Applies to both existing Tite and Mint nodes (see module docstring).
                ok1 = _enrich_existing_concept(c, "FRONT_TITE_FRONT", fields_common, source_rows)
                ok2 = _enrich_existing_concept(c, "FRONT_MINT_FRONT", fields_common, source_rows)
                if ok1 or ok2:
                    enriched += 1
            else:
                canonical_id = f"FRONT_{_slug(name)}"
                _upsert_concept(c, canonical_id, name, "DEFENSIVE_FRONTS", "Front family", fields_common, source_rows)
                imported += 1
        elif phase == "Formation-to-scheme":
            canonical_id = f"FILMSTUDY_{_slug(name)}"
            _upsert_concept(c, canonical_id, name, "FILM_STUDY", "Reading principle", fields_common, source_rows)
            imported += 1

    report["Formation Reading Lab"] = {
        "data_rows": len(rows), "imported": imported, "enriched_existing": enriched,
        "domain": "FORMATIONS / FILM_STUDY / DEFENSIVE_FRONTS / DEFENSIVE_PERSONNEL",
    }


# ============================================================================
# Defense Masterclass -- coverage rows enrich the existing coverage module;
# front/pressure rows merge with Scheme Concepts' front/pressure nodes;
# everything else (fit, movement, match rules, disguise, situation, film
# grade) becomes new FB_CONCEPT nodes.
# ============================================================================

DEFENSE_MASTERCLASS_COVERAGE_ENRICH = {
    "cover 0": "COVER_0", "cover 1": "COVER_1", "cover 2": "COVER_2", "tampa 2": "TAMPA_2",
    "cover 3 spot drop": "COVER_3", "cover 3 match": "COVER_3", "quarters": "COVER_4",
    "palms / 2-read": "PALMS_COVERAGE", "cover 6": "COVER_6", "bracket": "BRACKET_COVERAGE",
    "robber": "ROBBER_COVERAGE",
}
DEFENSE_MASTERCLASS_FRONT_ENRICH = {
    "over": "FRONT_4_3_OVER", "under": "FRONT_4_3_UNDER", "tite": "FRONT_TITE_FRONT",
    "mint": "FRONT_MINT_FRONT", "bear": "FRONT_BEAR_FRONT",
}
DEFENSE_MASTERCLASS_PRESSURE_ENRICH = {
    "sim pressure": "PRESSURE_SIMULATED_PRESSURE", "creeper": "PRESSURE_CREEPER_PRESSURE",
    "fire zone": "PRESSURE_FIRE_ZONE", "double mug": "PRESSURE_DOUBLE_MUG",
}
DEFENSE_MASTERCLASS_LAYER_DOMAIN = {
    "Front": "DEFENSIVE_FRONTS", "Movement": "DEFENSIVE_FRONTS", "Fit": "RUN_FITS",
    "Coverage": "COVERAGES", "Match rules": "COVERAGES", "Pressure": "PRESSURES",
    "Disguise": "COVERAGES", "Situation": "SITUATIONAL", "Film grade": "FILM_STUDY",
}
_DEFENSE_MASTERCLASS_MEG_NAMES = {"mod / meg / cone"}


def ingest_defense_masterclass(c, wb, report: dict):
    ws = wb["Defense Masterclass"]
    rows = _rows(ws)
    imported = 0
    enriched = 0
    for row_num, rec in rows:
        layer = rec.get("Layer")
        name = rec.get("Front / Coverage / Tool")
        if not name or not layer:
            continue
        name_key = name.strip().lower()
        fields = {
            "structure": rec.get("Structure"), "primary_purpose": rec.get("Primary Purpose"),
            "offensive_counter": rec.get("Offensive Counter"),
        }
        fields = {k: v for k, v in fields.items() if v and str(v).strip()}
        source_rows = [{"sheet": "Defense Masterclass", "row": row_num}]

        target = None
        if layer == "Coverage":
            target = DEFENSE_MASTERCLASS_COVERAGE_ENRICH.get(name_key)
        elif layer == "Front":
            target = DEFENSE_MASTERCLASS_FRONT_ENRICH.get(name_key)
        elif layer == "Pressure":
            target = DEFENSE_MASTERCLASS_PRESSURE_ENRICH.get(name_key)
        elif layer == "Match rules" and name_key in _DEFENSE_MASTERCLASS_MEG_NAMES:
            target = "MEG_TECHNIQUE"

        if target and _enrich_existing_concept(c, target, fields, source_rows):
            enriched += 1
            continue

        domain = DEFENSE_MASTERCLASS_LAYER_DOMAIN.get(layer, "DEFENSIVE_PHILOSOPHY")
        prefix = {
            "Front": "FRONT", "Movement": "FRONTTECH", "Fit": "RUNFIT", "Coverage": "COVERAGE",
            "Match rules": "COVERAGETECH", "Pressure": "PRESSURE", "Disguise": "COVERAGETECH",
            "Situation": "SITUATION", "Film grade": "FILMGRADE",
        }.get(layer, "DEFCONCEPT")
        canonical_id = f"{prefix}_{_slug(name)}"
        _upsert_concept(c, canonical_id, name, domain, layer, fields, source_rows)
        imported += 1

    report["Defense Masterclass"] = {
        "data_rows": len(rows), "imported": imported, "enriched_existing": enriched,
        "domain": "DEFENSIVE_FRONTS / RUN_FITS / COVERAGES / PRESSURES / SITUATIONAL / FILM_STUDY",
    }


# ============================================================================
# Scheme Taxonomy Guide / Defensive Taxonomy Guide -- meta "how this
# encyclopedia is organized" reference sheets. The workbook's own authors
# designed these as a schema guide (their "Recommended App Field" column
# literally names the payload keys this script independently arrived at,
# e.g. personnel_groupings/formation_architecture/run_foundation --
# confirming, not coincidence: this script's field-naming follows the
# source's own stated design intent). Routed per-Layer to the domain it
# actually documents; the two rows that are pure app-design meta (Evidence
# status, App rule) go to ENCYCLOPEDIA_GUIDE as a small "how this is
# organized / why projections are labeled" reference page.
# ============================================================================

OFFENSE_TAXONOMY_LAYER_DOMAIN = {
    "Family blend": "OFFENSIVE_SYSTEMS", "Personnel": "PERSONNEL", "Formation": "FORMATIONS",
    "Motion / shift": "FORMATIONS", "Run foundation": "RUN_GAME", "Run complements": "RUN_GAME",
    "Pass architecture": "PASSING_CONCEPTS", "RPO / option": "RUN_GAME",
    "Protection": "PASS_PROTECTION", "Tempo": "PLAY_CALLING", "Sequencing": "PLAY_CALLING",
    "Evidence status": "ENCYCLOPEDIA_GUIDE", "App rule": "ENCYCLOPEDIA_GUIDE",
}
DEFENSE_TAXONOMY_LAYER_DOMAIN = {
    "Family blend": "DEFENSIVE_PHILOSOPHY", "Personnel": "DEFENSIVE_PERSONNEL",
    "Front": "DEFENSIVE_FRONTS", "Technique": "DEFENSIVE_FRONTS", "Fit rule": "RUN_FITS",
    "Coverage shell": "COVERAGES", "Match rules": "COVERAGES", "Pressure": "PRESSURES",
    "Disguise": "COVERAGES", "Evidence status": "ENCYCLOPEDIA_GUIDE", "App rule": "ENCYCLOPEDIA_GUIDE",
}


def ingest_taxonomy_guides(c, wb, report: dict):
    for sheet_name, layer_domain, prefix in (
        ("Scheme Taxonomy Guide", OFFENSE_TAXONOMY_LAYER_DOMAIN, "OTAX"),
        ("Defensive Taxonomy Guide", DEFENSE_TAXONOMY_LAYER_DOMAIN, "DTAX"),
    ):
        ws = wb[sheet_name]
        rows = _rows(ws)
        imported = 0
        for row_num, rec in rows:
            layer = rec.get("Layer")
            if not layer:
                continue
            domain = layer_domain.get(layer, "ENCYCLOPEDIA_GUIDE")
            canonical_id = f"{prefix}_{_slug(layer)}"
            fields = {
                "question_it_answers": rec.get("Question It Answers"), "examples": rec.get("Examples"),
                "do_not_treat_as": rec.get("Do Not Treat As"),
                "recommended_app_field": rec.get("Recommended App Field"),
                "verification_method": rec.get("Verification Method"),
                "why_it_matters": rec.get("Why It Matters"),
            }
            fields = {k: v for k, v in fields.items() if v and str(v).strip()}
            _upsert_concept(c, canonical_id, f"{layer} (how {sheet_name.split(' Taxonomy')[0].lower()} concepts are organized)",
                             domain, layer, fields, [{"sheet": sheet_name, "row": row_num}])
            imported += 1
        report[sheet_name] = {"data_rows": len(rows), "imported": imported, "domain": "multiple (meta/taxonomy)"}


# ============================================================================
# NFL Scheme Profiles / FBS Scheme Profiles -- team-specific, NEEDS_
# VERIFICATION content (see module docstring). Every one of the 29 real
# source columns becomes a `fields` entry verbatim -- no fabrication, no
# collapsing to a single-label offense/defense name (per the mission's
# explicit "do not reduce teams to simplistic labels" instruction and the
# source's own "App rule" taxonomy guidance).
# ============================================================================

def ingest_scheme_profiles(c, wb, report: dict):
    for sheet_name, league, team_col in (
        ("NFL Scheme Profiles", "NFL", "Team"), ("FBS Scheme Profiles", "CFB", "Program"),
    ):
        ws = wb[sheet_name]
        rows = _rows(ws)
        imported = 0
        for row_num, rec in rows:
            team = rec.get(team_col)
            season = rec.get("Season")
            if not team or not season:
                continue
            canonical_id = f"SCHEME_PROFILE_{league}_{season}_{_slug(team)}"
            fields = {_snake(col): val for col, val in rec.items()
                      if col not in (team_col, "Season") and val is not None and str(val).strip()}
            node_id = _team_node_id(canonical_id)
            payload = {
                "domain": "NFL_SCHEMES" if league == "NFL" else "CFB_SCHEMES",
                "league": league, "team": team, "season": season, "fields": fields,
                "source_rows": [{"sheet": sheet_name, "row": row_num}],
            }
            c.execute(
                "INSERT INTO knowledge_nodes(node_id, node_type, canonical_id, label, competition_id, "
                "payload_json, verification_status) VALUES (?,?,?,?,?,?,?) "
                "ON CONFLICT(node_type, canonical_id) DO UPDATE SET "
                "label=excluded.label, payload_json=excluded.payload_json, verification_status=excluded.verification_status",
                (node_id, "FB_TEAM_SCHEME_PROFILE", canonical_id, f"{team} ({season})", league,
                 json.dumps(payload), LINEAGE_PROJECTED),
            )
            imported += 1
        report[sheet_name] = {
            "data_rows": len(rows), "imported": imported,
            "domain": "NFL_SCHEMES" if league == "NFL" else "CFB_SCHEMES",
            "verification_status": LINEAGE_PROJECTED,
        }


# ============================================================================
# Greatest Offenses/Defenses (NFL + CFB) -- real, source-cited statistical
# leader seasons. FB_HISTORICAL_RECORD, not FB_CONCEPT -- these are facts
# about a specific team-season, browsable as case studies (mission Section
# "GREAT OFFENSES & GREAT DEFENSES"), not teachable concepts.
# ============================================================================

def ingest_historical_records(c, wb, report: dict):
    configs = [
        ("NFL Greatest Offenses", "NFL", "offense", "Team"),
        ("NFL Greatest Defenses", "NFL", "defense", "Team"),
        ("CFB Greatest Offenses", "CFB", "offense", "Program"),
        ("CFB Greatest Defenses", "CFB", "defense", "Program"),
    ]
    for sheet_name, league, side, team_col in configs:
        ws = wb[sheet_name]
        rows = _rows(ws)
        imported = 0
        for row_num, rec in rows:
            team = rec.get(team_col)
            season = rec.get("Season")
            if not team or not season:
                continue
            canonical_id = f"GREAT_{side.upper()}_{league}_{season}_{_slug(team)}"
            fields = {_snake(col): val for col, val in rec.items()
                      if col not in (team_col, "Season") and val is not None and str(val).strip()}
            node_id = _hist_node_id(canonical_id)
            payload = {
                "domain": "GREAT_UNITS", "league": league, "side": side, "team": team, "season": season,
                "fields": fields, "source_rows": [{"sheet": sheet_name, "row": row_num}],
            }
            c.execute(
                "INSERT INTO knowledge_nodes(node_id, node_type, canonical_id, label, competition_id, "
                "payload_json, verification_status) VALUES (?,?,?,?,?,?,?) "
                "ON CONFLICT(node_type, canonical_id) DO UPDATE SET "
                "label=excluded.label, payload_json=excluded.payload_json, verification_status=excluded.verification_status",
                (node_id, "FB_HISTORICAL_RECORD", canonical_id, f"{season} {team}", league,
                 json.dumps(payload), SOURCE_BACKED),
            )
            imported += 1
        report[sheet_name] = {"data_rows": len(rows), "imported": imported, "domain": "GREAT_UNITS"}


# ============================================================================
# Film Charting Example -- real WORKED TEACHING EXAMPLES, explicitly
# labeled by the source itself (Game_ID = "EXAMPLE_NFL"/"EXAMPLE_CFB", never
# a real game id) as illustrative, not claims about any real game. Imported
# as-is with that labeling preserved verbatim -- never presented as "this
# really happened in game X," which the source deliberately does not claim.
# ============================================================================

def ingest_film_examples(c, wb, report: dict):
    ws = wb["Film Charting Example"]
    rows = _rows(ws)
    imported = 0
    for row_num, rec in rows:
        play_id = rec.get("Play_ID")
        game_id = rec.get("Game_ID")
        if not play_id or not game_id:
            continue
        canonical_id = f"FILMEX_{_slug(game_id)}_{_slug(play_id)}"
        fields = {_snake(col): val for col, val in rec.items()
                  if col not in ("Play_ID", "Game_ID") and val is not None and str(val).strip()}
        node_id = f"KN|FB_FILM_EXAMPLE|{canonical_id}"
        payload = {
            "domain": "FILM_STUDY", "is_illustrative_example": True, "game_id_label": game_id,
            "fields": fields, "source_rows": [{"sheet": "Film Charting Example", "row": row_num}],
        }
        c.execute(
            "INSERT INTO knowledge_nodes(node_id, node_type, canonical_id, label, competition_id, "
            "payload_json, verification_status) VALUES (?,?,?,?,?,?,?) "
            "ON CONFLICT(node_type, canonical_id) DO UPDATE SET "
            "label=excluded.label, payload_json=excluded.payload_json, verification_status=excluded.verification_status",
            (node_id, "FB_FILM_EXAMPLE", canonical_id, f"Worked example: {game_id} {play_id}", None,
             json.dumps(payload), SOURCE_BACKED),
        )
        imported += 1
    report["Film Charting Example"] = {"data_rows": len(rows), "imported": imported, "domain": "FILM_STUDY"}


# ============================================================================
# Relationships -- deliberately a modest, hand-verified set (quality over
# density): each edge below was checked against the actual source text of
# BOTH endpoints (not inferred from name similarity), matching the mission's
# "build these relationships only when football-correct and source-
# supported" instruction. This is a real starting graph, not an attempt at
# exhaustive coverage of 674 concepts -- a future pass can safely extend it
# using this same discipline.
# ============================================================================

ENCYCLOPEDIA_EDGES = [
    # Front lineage -- Formation Reading Lab's base fronts vs. Scheme
    # Concepts' strength-declared variants of them.
    ("FRONT_4_3_OVER", "VARIATION_OF", "FRONT_BASE_4_3"),
    ("FRONT_4_3_UNDER", "VARIATION_OF", "FRONT_BASE_4_3"),
    ("FRONT_ODD_FRONT", "RELATED_TO", "FRONT_BASE_3_4"),
    # "Tite... a three-down front with a zero-technique nose" is explicitly
    # framed as a member of the odd-front family (Scheme Concepts' own Odd
    # Front row: "often presented as 3-4, 3-3-5 or mint/tite structures").
    ("FRONT_TITE_FRONT", "VARIATION_OF", "FRONT_ODD_FRONT"),
    # "Mint Front... A Tite-like 3-3 structure" -- the source's own words.
    ("FRONT_MINT_FRONT", "VARIATION_OF", "FRONT_TITE_FRONT"),
    ("FRONT_WIDE_9", "USES_TECHNIQUE", "FRONT_ODD_FRONT"),  # Wide-9 is framed as an edge-technique choice layered onto a base front, not season/personnel-package specific

    # Defensive personnel packages this front/coverage vocabulary is
    # typically played out of.
    ("DEFPERSONNEL_NICKEL_4_2_5", "USES_SHELL", "COVER_4"),  # "Quarters/match... sound box fits" is the Nickel/4-2-5 row's own framing
    ("DEFPERSONNEL_3_3_5_STACK", "RELATED_TO", "FRONT_ODD_FRONT"),  # three-down, matches the odd-front interior picture

    # Run-concept family relationships -- gap vs. zone, and named
    # complements/counters of a base run.
    ("RUN_SPLIT_ZONE", "VARIATION_OF", "RUN_INSIDE_ZONE"),  # "Inside zone paired with a slicer" -- source's own definition
    ("RUN_OUTSIDE_ZONE_BOOT", "RELIES_ON", "RUN_WIDE_ZONE"),  # "Outside-zone action is paired with a quarterback keeper"
    ("RUN_MID_ZONE", "RELATED_TO", "RUN_WIDE_ZONE"),
    ("RUN_MID_ZONE", "RELATED_TO", "RUN_INSIDE_ZONE"),
    ("RUN_COUNTER_GT", "RELATED_TO", "RUN_POWER"),  # both gap-scheme pulling concepts, source lists them adjacently as the same "Gap run" family
    ("RUN_ZONE_READ", "VARIATION_OF", "RUN_INSIDE_ZONE"),  # "Zone blocking plus quarterback read of backside edge"
    ("RUN_POWER_READ", "RELATED_TO", "RUN_POWER"),  # "pairs sweep action with power blocking"
    ("RUN_TRIPLE_OPTION", "RELATED_TO", "RUN_INSIDE_VEER"),  # both dive-QB-pitch option structures
    ("RUN_INSIDE_VEER", "RELATED_TO", "RUN_MIDLINE"),
    ("RUN_ZONE_READ", "RELATED_TO", "RUN_SPEED_OPTION"),  # both are QB-perimeter-edge-read option concepts

    # Offensive systems -- what run/pass foundation each system is actually
    # built on, per the system's own "what it is" text.
    ("SYSTEM_SHANAHAN_KUBIAK_FAMILY", "RELIES_ON", "RUN_WIDE_ZONE"),
    ("SYSTEM_SHANAHAN_KUBIAK_FAMILY", "RELIES_ON", "RUN_MID_ZONE"),
    ("SYSTEM_SPREAD_OPTION", "RELIES_ON", "RUN_ZONE_READ"),
    ("SYSTEM_AIR_RAID", "USES_ROUTE", "PASSCONCEPT_MESH"),  # "Mesh, Four Verticals, Y-Cross, Y-Sail, Stick and screens" -- Air Raid's own concept menu
    ("SYSTEM_AIR_RAID", "USES_ROUTE", "PASSCONCEPT_FOUR_VERTICALS"),
    ("SYSTEM_AIR_RAID", "USES_ROUTE", "PASSCONCEPT_Y_CROSS"),
    ("SYSTEM_AIR_RAID", "USES_ROUTE", "PASSCONCEPT_STICK"),

    # Pass concepts -- Flood/Sail are distinct but closely related three-
    # level stretch concepts (see module docstring); Smash/Dagger both
    # attack two-high/single-high safety leverage voids the way Mills does.
    ("PASSCONCEPT_SAIL", "RELATED_TO", "PASSCONCEPT_FLOOD"),
    ("PASSCONCEPT_DAGGER", "RELATED_TO", "PASSCONCEPT_MILLS"),  # both a vertical clear-out opening a deep in-breaker
    ("PASSCONCEPT_SMASH", "CONTRASTS_WITH", "COVER_2"),  # "Efficient sideline answer against Cover 2" -- source's own framing
    ("PASSCONCEPT_FOUR_VERTICALS", "STRESSES", "COVER_3"),  # matches the coverage module's own existing STRESSES edge target direction/style
    ("PASSCONCEPT_MESH", "USES_ROUTE", "ROUTE_DRAG_SHALLOW"),  # "Two shallow crossers" == the Drag/Shallow route entry

    # Pressures -- Fire Zone and Simulated Pressure are both "show more,
    # rush four/five with a dropper" families; Double Mug is commonly the
    # pre-snap presentation that becomes one of them.
    ("PRESSURE_FIRE_ZONE", "RELATED_TO", "PRESSURE_SIMULATED_PRESSURE"),
    ("PRESSURE_DOUBLE_MUG", "RELATED_TO", "PRESSURE_CREEPER_PRESSURE"),

    # Defensive philosophy families -- what shell/front/coverage vocabulary
    # each named family actually plays, per its own row text.
    ("DEFPHILOSOPHY_FANGIO_FAMILY", "USES_SHELL", "TWO_HIGH_SHELL"),
    ("DEFPHILOSOPHY_FANGIO_FAMILY", "RELATED_TO", "FRONT_TITE_FRONT"),  # "adaptable odd/even fronts" -- Tite is the named odd-front variant this module has
    ("DEFPHILOSOPHY_SEATTLE_COVER_3_FAMILY", "USES_SHELL", "SINGLE_HIGH_SHELL"),
    ("DEFPHILOSOPHY_SEATTLE_COVER_3_FAMILY", "RELATED_TO", "COVER_3"),
    ("DEFPHILOSOPHY_SABAN_PATTERN_MATCH", "RELATED_TO", "MATCH_COVERAGE"),
]


# The 32 canonical_ids build_coverage_module.py originally created --
# predates this module's `domain` field convention, so they have none.
# Additive-only: sets `domain: "COVERAGES"` if and only if missing, never
# touches any field that module owns.
COVERAGE_MODULE_CANONICAL_IDS = [
    "MAN_COVERAGE", "ZONE_COVERAGE", "COVER_0", "COVER_1", "COVER_2", "TAMPA_2", "COVER_2_MAN",
    "COVER_3", "COVER_4", "COVER_6", "COVER_1_ROBBER", "COVER_2_INVERT", "COVER_3_CLOUD",
    "COVER_3_SKY", "COVER_3_BUZZ", "SINGLE_HIGH_SHELL", "TWO_HIGH_SHELL", "PRESS_TECHNIQUE",
    "OFF_TECHNIQUE", "TRAIL_TECHNIQUE", "BAIL_TECHNIQUE", "CATCH_TECHNIQUE", "LEVERAGE",
    "CUSHION", "MEG_TECHNIQUE", "MATCH_COVERAGE", "ROBBER_COVERAGE", "BRACKET_COVERAGE",
    "PALMS_COVERAGE", "BANJO_COVERAGE", "PREVENT_DEFENSE", "COVERAGE_DISGUISE",
]


# Meta/structural keys on the original coverage-module payload shape that
# should NOT be re-shown as a generic "field" (they're either already
# rendered specially, e.g. subcategory/verification_status, or are
# provenance, not content).
_COVERAGE_PAYLOAD_META_KEYS = {
    "concept_family", "source_rows", "source_sheet", "verification_status",
    "domain", "subcategory", "fields", "encyclopedia_fields", "encyclopedia_source_rows",
}


def tag_coverage_module_domain(c) -> int:
    """Additive-only: gives each of the 32 original coverage-module concepts
    a `domain`/`subcategory` (matching every other concept's convention,
    see module docstring) AND a `fields` dict built from ITS OWN existing
    top-level keys (summary/strengths/weaknesses/pre_snap_indicators/etc,
    all already real, human-authored content from build_coverage_module.py
    -- never invented here) merged with the `encyclopedia_fields` this
    module already added. This gives every FB_CONCEPT node ONE consistent
    shape for the frontend to render generically, without duplicating any
    content or touching a single field build_coverage_module.py owns."""
    tagged = 0
    for cid in COVERAGE_MODULE_CANONICAL_IDS:
        row = c.execute(
            "SELECT payload_json FROM knowledge_nodes WHERE node_type='FB_CONCEPT' AND canonical_id=?", (cid,)
        ).fetchone()
        if not row:
            continue
        payload = json.loads(row["payload_json"])
        if "domain" in payload and "fields" in payload:
            continue
        payload["domain"] = "COVERAGES"
        payload["subcategory"] = payload.get("concept_family")
        fields = {k: v for k, v in payload.items() if k not in _COVERAGE_PAYLOAD_META_KEYS and v is not None}
        fields.update(payload.get("encyclopedia_fields") or {})
        payload["fields"] = fields
        c.execute(
            "UPDATE knowledge_nodes SET payload_json=? WHERE node_type='FB_CONCEPT' AND canonical_id=?",
            (json.dumps(payload), cid),
        )
        tagged += 1
    return tagged


# ============================================================================
# Offensive personnel packages -- the standard RB-count/TE-count numbering
# convention (tens digit = running backs, ones digit = tight ends, the
# remainder of the 5 skill spots are wide receivers) is universal, fixed
# football terminology, not a team-specific or debatable claim -- safe to
# state directly without a per-row workbook citation, the same
# AUTHORED_FROM_ESTABLISHED_KNOWLEDGE discipline build_coverage_module.py
# already uses for widely-agreed football facts the source data implies but
# doesn't spell out row-by-row. The workbook's OWN Scheme Taxonomy Guide row
# (OTAX_PERSONNEL, imported above) already gives the real examples list
# ("10, 11, 12, 13, 20, 21, empty") this authoring is grounded in.
# ============================================================================
AUTHORED = "AUTHORED_FROM_ESTABLISHED_KNOWLEDGE"

PERSONNEL_PACKAGES = [
    ("00", 0, 0, "Empty backfield, no tight end -- five wide receivers. Rare; an extreme spread/passing-down look."),
    ("10", 1, 0, "1 running back, 0 tight ends, 4 wide receivers. A spread passing package."),
    ("11", 1, 1, "1 running back, 1 tight end, 3 wide receivers. The most common personnel grouping in modern football -- balances run/pass without tipping intent."),
    ("12", 1, 2, "1 running back, 2 tight ends, 2 wide receivers. Adds an extra in-line/H-back gap for the run game while keeping a real passing threat."),
    ("13", 1, 3, "1 running back, 3 tight ends, 1 wide receiver. A heavy run/play-action package, common in short-yardage and goal-line situations."),
    ("20", 2, 0, "2 running backs (often including a fullback), 0 tight ends, 3 wide receivers. An old-school two-back spread look."),
    ("21", 2, 1, "2 running backs (typically 1 RB + 1 FB), 1 tight end, 2 wide receivers. The traditional pro-style base package -- I-formation and offset-I both live here."),
    ("22", 2, 2, "2 running backs, 2 tight ends, 1 wide receiver. A heavy run-first package, common in short-yardage, goal-line and cold-weather run-heavy game plans."),
]


def ingest_personnel_packages(c) -> int:
    imported = 0
    for code, rb, te, desc in PERSONNEL_PACKAGES:
        wr = 5 - rb - te
        canonical_id = f"PERSONNEL_{code}"
        fields = {
            "running_backs": rb, "tight_ends": te, "wide_receivers": wr,
            "description": desc,
            "numbering_convention": "Tens digit = running backs on the field; ones digit = tight ends; "
                                     "the remaining skill spots (of 5 eligible non-lineman positions) are wide receivers.",
        }
        _upsert_concept(
            c, canonical_id, f"{code} Personnel", "PERSONNEL", "Offensive personnel packages", fields,
            [{"sheet": "Scheme Taxonomy Guide", "row": 3}],
            verification_status=AUTHORED,
        )
        imported += 1
    return imported


def ingest_relationships(c) -> int:
    written = 0
    for src, pred, tgt in ENCYCLOPEDIA_EDGES:
        src_row = c.execute("SELECT 1 FROM knowledge_nodes WHERE node_type='FB_CONCEPT' AND canonical_id=?", (src,)).fetchone()
        tgt_row = c.execute("SELECT 1 FROM knowledge_nodes WHERE node_type='FB_CONCEPT' AND canonical_id=?", (tgt,)).fetchone()
        if not src_row or not tgt_row:
            raise SystemExit(f"ABORT: relationship references a concept that doesn't exist: {src} -{pred}-> {tgt}")
        _add_edge(c, src, pred, tgt)
        written += 1
    return written


# ============================================================================
# Orchestration
# ============================================================================

def build(c) -> dict:
    _ensure_schema(c)
    wb = _load_wb()
    report: dict = {}
    ingest_simple_sheets(c, wb, report)
    ingest_route_tree(c, wb, report)
    ingest_scheme_concepts(c, wb, report)
    ingest_run_game_remainder(c, wb, report)
    ingest_passing_concepts_remainder(c, wb, report)
    ingest_trench_play(c, wb, report)
    ingest_formation_reading_lab(c, wb, report)
    ingest_defense_masterclass(c, wb, report)
    ingest_taxonomy_guides(c, wb, report)
    ingest_scheme_profiles(c, wb, report)
    ingest_historical_records(c, wb, report)
    ingest_film_examples(c, wb, report)
    personnel_packages_imported = ingest_personnel_packages(c)
    report["Personnel packages (authored, standard convention)"] = {
        "data_rows": 0, "imported": personnel_packages_imported, "domain": "PERSONNEL",
        "verification_status": AUTHORED,
    }
    coverage_domain_tagged = tag_coverage_module_domain(c)
    edges_written = ingest_relationships(c)
    c.commit()

    totals = {
        "sheets_processed": len(report),
        "knowledge_nodes_written": sum(v.get("imported", 0) for v in report.values()),
        "existing_coverage_nodes_enriched": sum(v.get("enriched_existing", 0) + v.get("enriched_existing_coverage", 0) for v in report.values()),
        "relationships_written": edges_written,
        "coverage_module_concepts_tagged_with_domain": coverage_domain_tagged,
    }
    return {"per_sheet": report, "totals": totals}


if __name__ == "__main__":
    conn = engine.connect()
    result = build(conn)
    conn.close()
    print(json.dumps(result, indent=2, default=str))
